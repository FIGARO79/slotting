"""
Router para la planificación de conteos de inventario.
Genera un archivo Excel con los conteos sugeridos basado en la clasificación ABC y el historial.
"""
import datetime
import random
from io import BytesIO
import polars as pl
import numpy as np
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.core.db import get_db
from app.models.schemas import CountExecutionRequest
from app.models.sql_models import CycleCount, CycleCountRecording, MasterItem
from app.services import csv_handler
from app.utils.auth import login_required, permission_required

import orjson
import os
from pydantic import BaseModel

router = APIRouter(prefix="/api/planner", tags=["planner"])

# --- Persistencia de Configuración ---
from app.core.config import PLANNER_CONFIG_PATH, PLANNER_DATA_PATH

CONFIG_FILE = PLANNER_CONFIG_PATH
PLAN_DATA_FILE = PLANNER_DATA_PATH

def load_config():
    """Carga la configuración desde el archivo JSON, o usa defaults."""
    default_holidays = [
        "2026-01-01", "2026-01-12", "2026-03-23", "2026-04-02", "2026-04-03",
        "2026-05-01", "2026-05-18", "2026-06-08", "2026-06-15", "2026-06-29",
        "2026-07-20", "2026-08-07", "2026-08-17", "2026-10-12", "2026-11-02",
        "2026-11-16", "2026-12-08", "2026-12-25"
    ]
    
    default_config = {
        "start_date": f"{datetime.datetime.now().year}-01-01",
        "end_date": f"{datetime.datetime.now().year}-12-31",
        "holidays": default_holidays
    }
    
    config = default_config
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'rb') as f:
                loaded = orjson.loads(f.read())
                config.update(loaded)
        except Exception:
            pass
            
    global HOLIDAYS
    try:
        HOLIDAYS = {datetime.datetime.strptime(d, "%Y-%m-%d").date() for d in config.get("holidays", [])}
    except ValueError:
        HOLIDAYS = set()
        
    return config

def save_config(config_data):
    """Guarda la configuración en el archivo JSON."""
    with open(CONFIG_FILE, 'wb') as f:
        f.write(orjson.dumps(config_data, option=orjson.OPT_INDENT_2))
        
    global HOLIDAYS
    try:
        HOLIDAYS = {datetime.datetime.strptime(d, "%Y-%m-%d").date() for d in config_data.get("holidays", [])}
    except ValueError:
        pass

def load_plan_data():
    """Carga los datos del plan calculeado/guardado."""
    if not os.path.exists(PLAN_DATA_FILE):
        return None
    try:
        with open(PLAN_DATA_FILE, 'rb') as f:
            return orjson.loads(f.read())
    except Exception:
        return None

def save_plan_data(data):
    """Guarda los datos del plan en JSON."""
    temp_file = f"{PLAN_DATA_FILE}.tmp"
    with open(temp_file, 'wb') as f:
        f.write(orjson.dumps(data, option=orjson.OPT_INDENT_2))
    os.replace(temp_file, PLAN_DATA_FILE)

# Cargar configuración inicial
PLANNER_CONFIG = load_config()

class PlannerConfigModel(BaseModel):
    start_date: str
    end_date: str
    holidays: list[str]

FREQUENCY_MAP = {'A': 3, 'B': 2, 'C': 1}

def get_working_days(start_date: datetime.date, end_date: datetime.date):
    """Genera una lista de días hábiles, excluyendo fines de semana y festivos configurados."""
    working_days = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5 and current_date not in HOLIDAYS:
            working_days.append(current_date)
        current_date += datetime.timedelta(days=1)
    return working_days

async def calculate_count_plan_data(start_date: str, end_date: str, db: AsyncSession):
    """Lógica central para calcular el plan de conteos distribuidos en días hábiles."""
    try:
        s_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        e_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD.")

    if s_date > e_date:
        raise HTTPException(status_code=400, detail="La fecha de inicio debe ser anterior a la fecha de fin.")

    stmt_master = select(MasterItem).where(MasterItem.physical_qty > 0)
    result_master = await db.execute(stmt_master)
    items_db = result_master.scalars().all()
    
    if not items_db:
        from app.services.csv_to_db import sync_master_csv_to_db
        await sync_master_csv_to_db(db)
        result_master = await db.execute(stmt_master)
        items_db = result_master.scalars().all()

    if not items_db:
         raise HTTPException(status_code=500, detail="El maestro de items está vacío.")

    data_list = [{'Item_Code': item.item_code, 'ABC_Code_stockroom': item.abc_code, 'Item_Description': item.description} for item in items_db]
    import polars as pl
    items_pl = pl.DataFrame(data_list)

    current_year = datetime.datetime.now().year
    start_of_year = f"{current_year}-01-01"
    query = (select(CycleCount.item_code, func.count(CycleCount.id).label("count")).where(CycleCount.timestamp >= start_of_year).group_by(CycleCount.item_code))
    result = await db.execute(query)
    previous_counts_map = {row.item_code: row.count for row in result.all()}

    tasks_to_schedule = []
    for row in items_pl.iter_rows(named=True):
        item_code, abc_code, description = row['Item_Code'], row['ABC_Code_stockroom'], row['Item_Description']
        required = FREQUENCY_MAP.get(abc_code, 0)
        done = previous_counts_map.get(item_code, 0)
        pending = max(0, required - done)
        for _ in range(pending):
            tasks_to_schedule.append({"Item Code": item_code, "ABC Code": abc_code, "Description": description})

    if not tasks_to_schedule:
        return pl.DataFrame({"Item Code": [], "ABC Code": [], "Description": [], "Planned Date": []})
    
    working_days = get_working_days(s_date, e_date)
    if not working_days:
        raise HTTPException(status_code=400, detail="No hay días hábiles en el rango seleccionado.")
        
    random.shuffle(tasks_to_schedule)
    planned_rows = []
    num_days = len(working_days)
    for i, task in enumerate(tasks_to_schedule):
        assigned_date = working_days[i % num_days]
        planned_rows.append({"Item Code": task["Item Code"], "ABC Code": task["ABC Code"], "Description": task["Description"], "Planned Date": assigned_date})
    df_output = pl.DataFrame(planned_rows)
    return df_output.sort(["Planned Date", "Item Code"])

@router.get("/preview_plan")
async def preview_count_plan(start_date: str = Query(...), end_date: str = Query(...), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    df_output = await calculate_count_plan_data(start_date, end_date, db)
    df_output = df_output.with_columns(pl.col("Planned Date").cast(pl.Utf8))
    
    # Resúmenes con Polars
    summary_by_date = df_output.group_by("Planned Date").len(name="count").to_dicts()
    summary_by_abc = df_output.group_by("ABC Code").len(name="count").to_dicts()
    
    return {
        "total_items": len(df_output),
        "summary_by_date": summary_by_date,
        "summary_by_abc": summary_by_abc,
        "details": df_output.to_dicts()
    }

@router.get("/generate_plan")
async def generate_count_plan(start_date: str = Query(...), end_date: str = Query(...), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    import openpyxl
    from openpyxl.utils import get_column_letter

    df_output = await calculate_count_plan_data(start_date, end_date, db)
    df_output = df_output.with_columns(pl.col("Planned Date").cast(pl.Utf8))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Planificacion'
    ws.append(df_output.columns)
    for row in df_output.iter_rows():
        ws.append(list(row))
    for i, col_name in enumerate(df_output.columns, start=1):
        col_data = df_output[col_name].cast(pl.Utf8, strict=False)
        max_len = max(col_data.str.len_chars().max() or 0, len(col_name)) + 2
        ws.column_dimensions[get_column_letter(i)].width = float(max_len)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename=plan_conteos_{start_date}.xlsx"})

@router.get("/config")
async def get_planner_config(username: str = Depends(permission_required("planner"))):
    return PLANNER_CONFIG

@router.post("/config")
async def update_planner_config(config: PlannerConfigModel, username: str = Depends(permission_required("planner"))):
    try:
        global PLANNER_CONFIG
        PLANNER_CONFIG = config.dict()
        save_config(PLANNER_CONFIG)
        return {"message": "Configuración guardada", "config": PLANNER_CONFIG}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/current_plan")
async def get_current_plan(username: str = Depends(permission_required("planner"))):
    return load_plan_data() or {}

@router.post("/update_plan")
async def update_count_plan(start_date: str = Query(...), end_date: str = Query(...), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    """
    Actualiza la planificación de forma incremental:
    Conserva los ítems ya programados en el pasado y regenera solo el futuro respetando feriados.
    """
    today = datetime.date.today()
    
    # 1. Cargar plan actual para extraer el pasado
    current_plan = load_plan_data()
    past_details = []
    if current_plan and "details" in current_plan:
        for it in current_plan["details"]:
            try:
                p_date = datetime.datetime.strptime(it.get("Planned Date"), "%Y-%m-%d").date()
                if p_date < today:
                    past_details.append(it)
            except (ValueError, TypeError):
                continue

    # 2. Generar nueva planificación para el futuro (mañana en adelante)
    tomorrow = today + datetime.timedelta(days=1)
    tomorrow_str = tomorrow.strftime('%Y-%m-%d')
    
    # Asegurarse de que el rango futuro sea válido
    try:
        e_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        if tomorrow > e_date:
            # Si el rango es inválido o ya pasó, solo guardamos el pasado preserve
            result_data = {
                "total_items": len(past_details),
                "details": past_details,
                "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
            }
            save_plan_data(result_data)
            return result_data
    except Exception:
        pass

    # Calcular lo nuevo desde mañana
    df_future = await calculate_count_plan_data(tomorrow_str, end_date, db)
    future_details = df_future.with_columns(pl.col("Planned Date").cast(pl.Utf8)).to_dicts()
    
    # 3. Combinar y guardar
    all_details = past_details + future_details
    result_data = {
        "total_items": len(all_details),
        "details": all_details,
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }
    
    save_plan_data(result_data)
    return result_data

@router.get("/execution/daily_items")
async def get_daily_items_for_execution(date: str = Query(...), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    # 1. Verificar siempre conteos previos en la base de datos para esta fecha
    res_prev = await db.execute(select(CycleCountRecording).where(CycleCountRecording.planned_date == date))
    prev_counts = res_prev.scalars().all()

    has_previous_counts = len(prev_counts) > 0
    items_with_diff_count = len([r for r in prev_counts if r.difference != 0])
    previous_count_total = len(prev_counts)

    # 2. Intentar cargar ítems planificados
    plan_data = load_plan_data()
    daily_items = []

    if plan_data and "details" in plan_data:
        daily_items = [item for item in plan_data["details"] if item.get("Planned Date") == date]

    # Si no hay plan para hoy y tampoco hay conteos previos, retornar vacío rápido
    if not daily_items and not has_previous_counts:
        return {
            "items": [], 
            "has_previous_counts": False,
            "previous_count": 0,
            "items_with_diff_count": 0
        }

    item_codes = [item.get("Item Code") for item in daily_items]

    # Obtener info maestra para los ítems del plan
    res_master = await db.execute(select(MasterItem).where(MasterItem.item_code.in_(item_codes)))
    master_map = {m.item_code: m for m in res_master.scalars().all()}

    enriched = []
    for item in daily_items:
        m = master_map.get(item.get("Item Code"))
        enriched.append({
            "item_code": item.get("Item Code"),
            "description": item.get("Description"),
            "abc_code": item.get("ABC Code"),
            "bin_location": m.bin_1 if m else "N/A",
            "additional_locations": m.additional_bin if m and m.additional_bin else "",
            "system_qty": m.physical_qty if m else 0,
            "planned_date": date
        })

    return {
        "items": sorted(enriched, key=lambda x: x["bin_location"]),
        "has_previous_counts": has_previous_counts,
        "previous_count": previous_count_total,
        "items_with_diff_count": items_with_diff_count
    }
@router.get("/execution/items_with_differences")
async def get_items_with_differences(date: str = Query(...), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    """Carga solo los ítems que tuvieron diferencias en conteos previos para reconteo."""
    res_prev = await db.execute(select(CycleCountRecording).where(
        CycleCountRecording.planned_date == date,
        CycleCountRecording.difference != 0
    ))
    prev_items = res_prev.scalars().all()
    
    if not prev_items:
        return {"items": [], "total_items_with_diff": 0}
        
    item_codes = [r.item_code for r in prev_items]
    res_master = await db.execute(select(MasterItem).where(MasterItem.item_code.in_(item_codes)))
    master_map = {m.item_code: m for m in res_master.scalars().all()}
    
    enriched = []
    for prev in prev_items:
        m = master_map.get(prev.item_code)
        enriched.append({
            "item_code": prev.item_code,
            "description": prev.item_description,
            "abc_code": prev.abc_code or (m.abc_code if m else "C"),
            "bin_location": m.bin_1 if m else (prev.bin_location or "N/A"),
            "additional_locations": m.additional_bin if m and m.additional_bin else "",
            "system_qty": m.physical_qty if m else prev.system_qty,
            "planned_date": date
        })
        
    return {
        "items": sorted(enriched, key=lambda x: x["bin_location"]),
        "total_items_with_diff": len(enriched)
    }

@router.post("/execution/save")
async def save_daily_execution(execution_data: CountExecutionRequest, username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    """Guarda conteos con validación estricta de system_qty desde el servidor."""
    try:
        today_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
        item_codes = [it.item_code for it in execution_data.items]
        res_master = await db.execute(select(MasterItem).where(MasterItem.item_code.in_(item_codes)))
        master_map = {m.item_code: m for m in res_master.scalars().all()}

        saved, updated = 0, 0
        for item in execution_data.items:
            m_item = master_map.get(item.item_code)
            if not m_item: continue

            physical = item.physical_qty
            system = m_item.physical_qty or 0
            
            res_exist = await db.execute(select(CycleCountRecording).where(CycleCountRecording.item_code == item.item_code, CycleCountRecording.planned_date == execution_data.date))
            existing = res_exist.scalar_one_or_none()
            
            if existing:
                existing.physical_qty, existing.system_qty = physical, system
                existing.difference, existing.username, existing.executed_date = physical - system, username, today_iso
                updated += 1
            else:
                bin_loc = m_item.bin_1
                if m_item.additional_bin:
                    bin_loc = f"{bin_loc} | {m_item.additional_bin}"
                
                db.add(CycleCountRecording(
                    planned_date=execution_data.date, 
                    executed_date=today_iso, 
                    item_code=item.item_code, 
                    item_description=m_item.description, 
                    bin_location=bin_loc, 
                    system_qty=system, 
                    physical_qty=physical, 
                    difference=physical-system, 
                    username=username, 
                    abc_code=m_item.abc_code
                ))
                saved += 1


        await db.commit()
        return {"message": f"Guardados: {saved} nuevos, {updated} actualizados."}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/execution/stats")
async def get_execution_stats(year: int = Query(datetime.datetime.now().year), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(CycleCountRecording).where(CycleCountRecording.executed_date.like(f"{year}-%")))
    records = res.scalars().all()
    exec_grid = {cat: [0]*12 for cat in ['A', 'B', 'C']}
    delta_grid = {cat: [0]*12 for cat in ['A', 'B', 'C']}
    for r in records:
        try:
            m_idx = datetime.datetime.strptime(r.executed_date[:10], "%Y-%m-%d").month - 1
            cat = r.abc_code if r.abc_code in exec_grid else 'C'
            exec_grid[cat][m_idx] += 1
            if r.difference != 0: delta_grid[cat][m_idx] += 1
        except: continue
    return {"executed": exec_grid, "delta": delta_grid, "year": year}

@router.get('/cycle_count_differences')
async def get_cycle_count_differences(year: int = Query(None), month: int = Query(None), only_differences: bool = Query(True), username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    query = select(CycleCountRecording)
    if only_differences: query = query.where(CycleCountRecording.difference != 0)
    if year: query = query.where(CycleCountRecording.executed_date.like(f"{year}-%"))
    if month:
        m_str = str(month).zfill(2)
        query = query.where(CycleCountRecording.executed_date.like(f"%-%{m_str}-%"))
    res = await db.execute(query.order_by(CycleCountRecording.executed_date.desc()))
    return res.scalars().all()

@router.put('/cycle_count_differences/{rec_id}')
async def update_cycle_count_diff(rec_id: int, data: dict, username: str = Depends(permission_required("planner")), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(CycleCountRecording).where(CycleCountRecording.id == rec_id))
    r = res.scalar_one_or_none()
    if not r: raise HTTPException(status_code=404, detail="No encontrado")
    r.physical_qty = data['physical_qty']
    r.difference = r.physical_qty - r.system_qty
    await db.commit()
    return {"message": "Actualizado"}
